# -*- coding: utf-8 -*-
"""
============================================================================
 ANÁLISIS DE SENSIBILIDAD — PARÁMETROS DE COMPORTAMIENTO (Grupo B)
 TP Simulación — lockers última milla
============================================================================
 Acota el impacto de los parámetros que NO son observables, porque la red de
 lockers analizada todavía no está desplegada (estudio ex-ante):

   · Tiempo de retiro del locker      (±30 % sobre la mediana)
   · TMP, tiempo máximo de permanencia (±30 %)
   · Tasa de devolución                (±30 %)

 A diferencia del Grupo A (costos), acá SÍ cambia la dinámica del sistema,
 por lo que cada variante se vuelve a simular.

 Además re-optimiza la grilla bajo el parámetro más influyente (tiempo de
 retiro) para verificar si el diseño óptimo se desplaza.

 Genera: sensibilidad_supuestos.xlsx, fig6_sensibilidad_supuestos.png

 Corrida rápida:
   SENSB_TF_DIAS=30 SENSB_NREP=12 SENSB_NREP_GRID=6 python3 sensibilidad_b.py
============================================================================
"""
import os
import statistics as st
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

import simulacion_lockers as S

TF_DIAS = int(os.environ.get('SENSB_TF_DIAS', 0))
if TF_DIAS:
    S.TF = 60 * 24 * TF_DIAS
N_REP = int(os.environ.get('SENSB_NREP', 30))
N_REP_GRID = int(os.environ.get('SENSB_NREP_GRID', 8))
SEED0 = 1000
DELTA = 0.30

GRID_CLTC = [10, 20, 30, 40, 50, 60]
GRID_CLTM = [5, 10, 15, 20, 25]
GRID_CLTG = [5, 10, 15]
ESCENARIOS = ['NORMAL', 'NAVIDAD', 'CYBER']

RET_BASE = list(S.RETIROS)
DEV_BASE = list(S.DEVOLUCIONES)
TMP_BASE = S.TMP


def reset():
    S.RETIROS = list(RET_BASE)
    S.DEVOLUCIONES = list(DEV_BASE)
    S.TMP = TMP_BASE


def aplicar(param, nivel):
    """nivel in (-1, 0, +1) -> -30 %, base, +30 %."""
    reset()
    if nivel == 0:
        return
    f = 1 + DELTA * nivel
    if param == 'retiro':                       # mediana de retiro
        S.RETIROS = [x * f for x in RET_BASE]
    elif param == 'tmp':
        S.TMP = TMP_BASE * f
    elif param == 'devolucion':                 # TASA de devolución (intervalo = 1/tasa)
        S.DEVOLUCIONES = [x / f for x in DEV_BASE]


def optimos_desde_barrido():
    try:
        ws = load_workbook('resultados_barrido.xlsx', read_only=True)['Barrido']
        return {r[0]: (r[1], r[2], r[3]) for r in ws.iter_rows(min_row=2, values_only=True)
                if r[12] == 'SI'}
    except Exception as e:
        print(f"  (no se pudo leer resultados_barrido.xlsx: {e})")
        return {}


def medir(esc, cfg, nrep=N_REP):
    rr = [S.simular(esc, cfg, SEED0 + k) for k in range(nrep)]
    return {k: st.mean(r[k] for r in rr) for k in ('CPE', 'EF_pct', 'PPV', 'CI')}


opt = optimos_desde_barrido()
if not opt:
    d = S.DEFAULT_CONFIG
    opt = {e: (d['CLTC'], d['CLTM'], d['CLTG']) for e in ESCENARIOS}

PARAMS = [('retiro', 'Tiempo de retiro'),
          ('tmp', 'TMP'),
          ('devolucion', 'Tasa de devolución')]
NIVELES = [(-1, '-30 %'), (0, 'base'), (1, '+30 %')]

print(f"Sensibilidad de parámetros de comportamiento (±{DELTA:.0%}) | "
      f"TF={S.TF/1440:.0f}d | réplicas={N_REP}\n")

res = {}   # (esc, param, nivel) -> métricas
for esc in ESCENARIOS:
    c, m, g = opt[esc]
    cfg = dict(CLTC=c, CLTM=m, CLTG=g)
    print(f"===== {esc} (config {c}/{m}/{g}) =====")
    print(f"{'Parámetro':<22}{'Nivel':<9}{'CPE':>8}{'Sat.%':>9}{'PPV%':>8}{'CI':>7}")
    for pk, plabel in PARAMS:
        for niv, nlabel in NIVELES:
            aplicar(pk, niv)
            r = medir(esc, cfg)
            res[(esc, pk, niv)] = r
            print(f"{plabel:<22}{nlabel:<9}{r['CPE']:>8.2f}{r['EF_pct']:>9.2f}"
                  f"{r['PPV']:>8.2f}{r['CI']:>7.0f}")
    print()
reset()

# ------------------------------------------------------------------
# Re-optimización bajo el parámetro más influyente (tiempo de retiro)
# ------------------------------------------------------------------
print("¿Se desplaza el diseño óptimo al variar el tiempo de retiro? (escenario CYBER)\n")
reopt = {}
for niv, nlabel in NIVELES:
    aplicar('retiro', niv)
    mejor, mejor_cpe = None, float('inf')
    for cc in GRID_CLTC:
        for mm in GRID_CLTM:
            for gg in GRID_CLTG:
                cfg = dict(CLTC=cc, CLTM=mm, CLTG=gg)
                v = st.mean(S.simular('CYBER', cfg, SEED0 + k)['CPE']
                            for k in range(N_REP_GRID))
                if v < mejor_cpe:
                    mejor, mejor_cpe = (cc, mm, gg), v
    reopt[niv] = (mejor, mejor_cpe)
    print(f"  retiro {nlabel:<7} -> óptimo {mejor[0]}/{mejor[1]}/{mejor[2]}  CPE={mejor_cpe:.2f}")
reset()
print()

# ------------------------------------------------------------------ XLSX
wb = Workbook()
ws = wb.active; ws.title = "Sensibilidad_B"
ws.append(["escenario", "parametro", "nivel", "CPE", "saturacion_%", "PPV_%", "CI"])
for (esc, pk, niv), r in res.items():
    lbl = dict(NIVELES)[niv]
    nombre = dict(PARAMS)[pk]
    ws.append([esc, nombre, lbl, round(r['CPE'], 3), round(r['EF_pct'], 2),
               round(r['PPV'], 3), round(r['CI'], 0)])
ws2 = wb.create_sheet("Reoptimizacion")
ws2.append(["parametro", "nivel", "optimo", "CPE"])
for niv, lbl in NIVELES:
    cfg, v = reopt[niv]
    ws2.append(["Tiempo de retiro", lbl, f"{cfg[0]}/{cfg[1]}/{cfg[2]}", round(v, 3)])
wb.save("sensibilidad_supuestos.xlsx")

# ------------------------------------------------------------------ FIGURA
plt.rcParams.update({'font.size': 7, 'axes.titlesize': 8, 'axes.labelsize': 7,
                     'xtick.labelsize': 6.5, 'ytick.labelsize': 6.5,
                     'legend.fontsize': 6.5, 'figure.dpi': 200})
GREYS = ['0.15', '0.45', '0.70']; MARKS = ['o', 's', '^']
xs = [-1, 0, 1]; xt = ['−30 %', 'base', '+30 %']
fig, ax = plt.subplots(3, 1, figsize=(3.35, 4.6))
paneles = [('retiro', 'CPE', '(a) Tiempo de retiro', 'CPE (USD/paq)'),
           ('tmp', 'PPV', '(b) TMP', 'Paquetes vencidos (%)'),
           ('devolucion', 'CI', '(c) Tasa de devolución', 'Clientes insatisf.')]
for a, (pk, metrica, titulo, ylab) in zip(ax, paneles):
    for esc, cgrey, mk in zip(ESCENARIOS, GREYS, MARKS):
        ys = [res[(esc, pk, n)][metrica] for n in xs]
        a.plot(xs, ys, marker=mk, color=cgrey, lw=1.3, ms=3.5, label=esc.capitalize())
    a.set_xticks(xs); a.set_xticklabels(xt)
    a.set_title(titulo); a.set_ylabel(ylab)
    a.grid(alpha=0.25, lw=0.4)
ax[0].legend(frameon=False)
plt.tight_layout(); plt.savefig("fig6_sensibilidad_supuestos.png"); plt.close()

print("OK -> sensibilidad_supuestos.xlsx, fig6_sensibilidad_supuestos.png")
