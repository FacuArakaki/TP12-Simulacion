# -*- coding: utf-8 -*-
"""
Figuras para el PAPER (formato 2 columnas, A4).
Relee los .xlsx de resultados ya calculados -> NO vuelve a simular.
Estilo sobrio, apto impresión monocromática.
Genera: fig1_distribuciones.png .. fig5_tornado.png
"""
import os
from collections import defaultdict
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook

D = os.path.dirname(os.path.abspath(__file__))
plt.rcParams.update({'font.size': 7, 'axes.titlesize': 8, 'axes.labelsize': 7,
                     'xtick.labelsize': 6.5, 'ytick.labelsize': 6.5,
                     'legend.fontsize': 6.5, 'figure.dpi': 200})
W = 3.35          # ancho de columna en pulgadas
GREYS = ['0.15', '0.45', '0.70']
MARKS = ['o', 's', '^']

def col(f, hoja='Datos'):
    ws = load_workbook(os.path.join(D, f), read_only=True)[hoja]
    return [r[0] for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]

# ---------------------------------------------------------------- FIG 1
# Distribuciones de entrada: llegadas (3 escenarios), retiros, devoluciones
lleg = {e: col(f'dataset_llegadas_{e}.xlsx') for e in ('normal', 'navidad', 'cyber')}
ret = col('dataset_retiros.xlsx')
dev = col('dataset_devoluciones.xlsx')

fig, ax = plt.subplots(3, 1, figsize=(W, 4.6))
for (e, d), c, m in zip(lleg.items(), GREYS, MARKS):
    dd = [x for x in d if x < 25]
    ax[0].hist(dd, bins=25, histtype='step', linewidth=1.2, color=c,
               density=True, label=f"{e} ({sum(d)/len(d):.1f} min)")
ax[0].set_title('(a) Inter-arribo de entregas fallidas [observado]')
ax[0].set_xlabel('minutos'); ax[0].set_ylabel('densidad'); ax[0].legend(frameon=False)

ax[1].hist([x for x in ret if x < 4000], bins=40, color='0.55', edgecolor='0.2', linewidth=0.3)
ax[1].set_title('(b) Tiempo de retiro del locker [parametrizado]')
ax[1].set_xlabel('minutos'); ax[1].set_ylabel('frecuencia')

ax[2].hist(dev, bins=40, color='0.75', edgecolor='0.2', linewidth=0.3)
ax[2].set_title('(c) Inter-arribo de devoluciones [parametrizado]')
ax[2].set_xlabel('minutos'); ax[2].set_ylabel('frecuencia')
plt.tight_layout(); plt.savefig(os.path.join(D, 'fig1_distribuciones.png')); plt.close()

# ---------------------------------------------------------------- datos barrido
ws = load_workbook(os.path.join(D, 'resultados_barrido.xlsx'), read_only=True)['Barrido']
rows = [r for r in ws.iter_rows(min_row=2, values_only=True)]
# esc -> (c,m,g) -> cpe   /  esc -> optimo
cpe = defaultdict(dict); opt = {}
for r in rows:
    esc, c, m, g, _tot, v = r[0], r[1], r[2], r[3], r[4], r[5]
    cpe[esc][(c, m, g)] = v
    if r[12] == 'SI':
        opt[esc] = (c, m, g)
GC = sorted({k[0] for k in cpe['CYBER']}); GM = sorted({k[1] for k in cpe['CYBER']})
GG = sorted({k[2] for k in cpe['CYBER']})

# ---------------------------------------------------------------- FIG 2  curva U
fig, ax = plt.subplots(figsize=(W, 2.5))
for (esc, c, m) in zip(['NORMAL', 'NAVIDAD', 'CYBER'], GREYS, MARKS):
    _, bm, bg = opt[esc]
    xs = [x for x in GC if (x, bm, bg) in cpe[esc]]
    ys = [cpe[esc][(x, bm, bg)] for x in xs]
    ax.plot(xs, ys, marker=m, color=c, lw=1.3, ms=3.5, label=esc.capitalize())
ax.set_xlabel('Compartimentos chicos (CLTC)')
ax.set_ylabel('CPE (USD/paquete)')
ax.legend(frameon=False); ax.grid(alpha=0.25, lw=0.4)
plt.tight_layout(); plt.savefig(os.path.join(D, 'fig2_curvaU.png')); plt.close()

# ---------------------------------------------------------------- FIG 3  heatmaps (3 pares)
esc = 'CYBER'
bc, bm, bg = opt[esc]
GRIDS = {'CLTC': GC, 'CLTM': GM, 'CLTG': GG}
LAB = {'CLTC': 'chicos', 'CLTM': 'medianos', 'CLTG': 'grandes'}
OPTV = {'CLTC': bc, 'CLTM': bm, 'CLTG': bg}
pares = [('CLTC', 'CLTM'), ('CLTC', 'CLTG'), ('CLTM', 'CLTG')]
letra = ['(a)', '(b)', '(c)']

fig, axs = plt.subplots(3, 1, figsize=(W, 5.4))
for ax, (xk, yk), lt in zip(axs, pares, letra):
    fk = ({'CLTC', 'CLTM', 'CLTG'} - {xk, yk}).pop()
    xv, yv = GRIDS[xk], GRIDS[yk]
    mat = []
    for y in yv:
        fila = []
        for x in xv:
            cfg = {xk: x, yk: y, fk: OPTV[fk]}
            fila.append(cpe[esc].get((cfg['CLTC'], cfg['CLTM'], cfg['CLTG'])))
        mat.append(fila)
    im = ax.imshow(mat, cmap='Greys', aspect='auto', origin='lower')
    ax.set_xticks(range(len(xv))); ax.set_xticklabels(xv)
    ax.set_yticks(range(len(yv))); ax.set_yticklabels(yv)
    ax.set_xlabel(f'Compartimentos {LAB[xk]}')
    ax.set_ylabel(f'Comp. {LAB[yk]}')
    ax.set_title(f'{lt} {LAB[xk]} vs {LAB[yk]}  ({LAB[fk]} = {OPTV[fk]})')
    vmax = max(v for fila in mat for v in fila if v is not None)
    for i in range(len(yv)):
        for j in range(len(xv)):
            v = mat[i][j]
            if v is None: continue
            ax.text(j, i, f"{v:.1f}", ha='center', va='center', fontsize=5.2,
                    color='white' if v > 0.55 * vmax else 'black')
    cb = fig.colorbar(im, ax=ax, fraction=0.040, pad=0.02)
    cb.ax.tick_params(labelsize=5.5)
plt.tight_layout(); plt.savefig(os.path.join(D, 'fig3_heatmap.png')); plt.close()

# ---------------------------------------------------------------- FIG 4  comparación
ws = load_workbook(os.path.join(D, 'comparacion_configs.xlsx'), read_only=True)['Comparacion']
comp = defaultdict(dict)
for r in ws.iter_rows(min_row=2, values_only=True):
    comp[r[0]][r[1]] = (r[6], r[9])          # categoria -> (CPE, EF%)
CATS = ['Malo', 'Normal', 'Óptimo', 'Óptimo excelente']
fig, ax = plt.subplots(figsize=(W, 2.5))
xs = range(len(CATS)); wbar = 0.26
for i, (esc, c) in enumerate(zip(['NORMAL', 'NAVIDAD', 'CYBER'], GREYS)):
    ys = [comp[esc][k][0] for k in CATS]
    ax.bar([x + (i - 1) * wbar for x in xs], ys, wbar, color=c, label=esc.capitalize(),
           edgecolor='black', linewidth=0.4)
ax.set_yscale('log')
ax.set_xticks(list(xs)); ax.set_xticklabels(['Sub-\ndimens.', 'Sin\noptimizar', 'Óptimo', 'Máximo\nservicio'])
ax.set_ylabel('CPE (USD/paquete, log)')
ax.legend(frameon=False); ax.grid(axis='y', alpha=0.25, lw=0.4)
plt.tight_layout(); plt.savefig(os.path.join(D, 'fig4_comparacion.png')); plt.close()

# ---------------------------------------------------------------- FIG 5  tornado
ws = load_workbook(os.path.join(D, 'sensibilidad.xlsx'), read_only=True)['Tornado']
tor = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] == 'CYBER']
tor.sort(key=lambda r: abs(r[4] - r[3]))
base = tor[0][2]
fig, ax = plt.subplots(figsize=(W, 2.3))
labels = [r[1].split(' (')[0] for r in tor]
for y, r in enumerate(tor):
    ax.barh(y, r[3] - r[2], left=r[2], color='0.75', edgecolor='black', linewidth=0.4)
    ax.barh(y, r[4] - r[2], left=r[2], color='0.35', edgecolor='black', linewidth=0.4)
ax.axvline(base, color='black', lw=0.9, ls='--')
ax.set_yticks(range(len(tor))); ax.set_yticklabels(labels)
ax.set_xlabel('CPE (USD/paquete)')
ax.legend(handles=[plt.Rectangle((0, 0), 1, 1, fc='0.75', ec='black', lw=0.4),
                   plt.Rectangle((0, 0), 1, 1, fc='0.35', ec='black', lw=0.4)],
          labels=['−30 %', '+30 %'], frameon=False, loc='lower right')
ax.grid(axis='x', alpha=0.25, lw=0.4)
plt.tight_layout(); plt.savefig(os.path.join(D, 'fig5_tornado.png')); plt.close()

print("OK -> fig1..fig5 generadas")
for e in ('NORMAL', 'NAVIDAD', 'CYBER'):
    print(f"  óptimo {e}: {opt[e]}  CPE={cpe[e][opt[e]]:.2f}")
