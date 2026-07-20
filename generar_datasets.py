# -*- coding: utf-8 -*-
"""
Generador de datasets para TP Simulación - Logística Última Milla (Lockers).

REALES  (extraídos de los 4 archivos 'Semana x' de operación):
  - dataset_llegadas_{escenario}.xlsx : inter-arribos (min) de fallos por Cliente Ausente.
  - dataset_tamanos.xlsx              : distribución de tamaños por escenario.

SUPUESTOS (no existen en una operación sin lockers -> parametrizados y declarados):
  - dataset_retiros.xlsx      : tiempo de retiro del locker (min), lognormal cola derecha.
  - dataset_devoluciones.xlsx : inter-arribo de devoluciones (min), anclado a tasa NRF.
"""
import csv, os
from datetime import datetime
from collections import defaultdict
import numpy as np
import scipy.stats as stats
from openpyxl import Workbook

np.random.seed(42)  # reproducibilidad

SRC = "/sessions/vibrant-compassionate-hamilton/mnt/uploads"
OUT = "/sessions/vibrant-compassionate-hamilton/mnt/outputs"
os.makedirs(OUT, exist_ok=True)

I_ORD, I_EST, I_MOT, I_HORA, I_CAR = 8, 28, 29, 30, 59
FILES = [("Semana random.csv", "NORMAL"),
         ("Semana navidad.csv", "NAVIDAD"),
         ("Semana 1 cyber (2).csv", "CYBER"),
         ("Semana 2 cyber.csv", "CYBER")]

SIZES = {"MINI", "PAQUETERIA", "MEDIUM", "BIG-SIZE", "OVER-SIZE",
         "ALTO-MAXIMO", "FLEX", "TIENDA", "DEVOLUCIONES"}
ZONES = {"caba", "sur", "norte", "oeste", "la plata"}

# Mapeo A (COMPLETO): todo item mapeado a un tier del modelo.
TIER_FULL = {"MINI": "chico", "PAQUETERIA": "chico", "FLEX": "chico",
             "MEDIUM": "mediano",
             "BIG-SIZE": "grande", "OVER-SIZE": "grande", "ALTO-MAXIMO": "grande"}
# Mapeo B (LOCKEABLE): excluye muebles/voluminosos que NO caben en un locker.
#   BIG-SIZE (mesa ratona, vajillero, aire) y ALTO-MAXIMO no son lockeables.
TIER_LOCK = {"MINI": "chico", "PAQUETERIA": "chico", "FLEX": "chico",
             "MEDIUM": "mediano",
             "OVER-SIZE": "grande"}
# TIENDA / DEVOLUCIONES se excluyen siempre.


def parse_dt(s):
    s = s.strip()
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def split_car(s):
    size = None
    for tok in s.split("."):
        t = tok.strip()
        if t in SIZES:
            size = t
    return size


# ============================================================
# 1. LECTURA DE CRUDOS -> eventos de fallo por ausencia
# ============================================================
# scen -> lista de (datetime, orden); scen -> Counter de tamaño crudo
events = defaultdict(list)
raw_size = defaultdict(lambda: defaultdict(int))   # scen -> size cruda -> n

for fn, lab in FILES:
    with open(os.path.join(SRC, fn), encoding="utf-8", errors="replace") as f:
        r = csv.reader(f, delimiter=";")
        next(r)
        for row in r:
            if len(row) <= I_CAR:
                continue
            if row[I_MOT].strip() != "Cliente Ausente":
                continue
            dt = parse_dt(row[I_HORA])
            if dt is None:
                continue
            events[lab].append((dt, row[I_ORD]))
            sz = split_car(row[I_CAR])
            if sz:
                raw_size[lab][sz] += 1


def tier_probs(counts, mapping):
    agg = defaultdict(int)
    for sz, n in counts.items():
        t = mapping.get(sz)
        if t:
            agg[t] += n
    tot = agg["chico"] + agg["mediano"] + agg["grande"]
    if tot == 0:
        return (0, 0, 0), agg, 0
    return (agg["chico"] / tot, agg["mediano"] / tot, agg["grande"] / tot), agg, tot


# ============================================================
# 2. INTER-ARRIBOS INTRA-DÍA (min) por escenario  [REAL]
# ============================================================
def interarrivals(evs):
    # una visita fallida = un arribo; dedupe (orden, timestamp)
    uniq = sorted(set((dt, o) for dt, o in evs))
    by_day = defaultdict(list)
    for dt, o in uniq:
        by_day[dt.date()].append(dt)
    gaps = []
    for day, times in by_day.items():
        times.sort()
        for a, b in zip(times, times[1:]):
            g = (b - a).total_seconds() / 60.0
            if g >= 0:            # mismo día, sin saltos nocturnos
                gaps.append(g)
    return np.array(gaps, dtype=float)


llegadas = {sc: interarrivals(evs) for sc, evs in events.items()}


def hourly_counts(evs):
    """Conteo de arribos por (fecha, hora) en el horario operativo observado."""
    uniq = sorted(set((dt, o) for dt, o in evs))
    cnt = defaultdict(int)
    hours_seen = set()
    for dt, o in uniq:
        cnt[(dt.date(), dt.hour)] += 1
        hours_seen.add(dt.hour)
    lo, hi = min(hours_seen), max(hours_seen)
    days = sorted(set(d for d, h in cnt))
    counts = [cnt.get((d, h), 0) for d in days for h in range(lo, hi + 1)]
    return np.array(counts), (lo, hi)


def poisson_chi2(counts):
    """Bondad de ajuste Poisson sobre conteos por hora (agrupado)."""
    lam = counts.mean()
    kmax = int(counts.max())
    obs = np.bincount(counts, minlength=kmax + 1).astype(float)
    exp = stats.poisson.pmf(np.arange(kmax + 1), lam) * len(counts)
    # agrupar colas para exp>=5
    o, e = [], []
    acc_o = acc_e = 0.0
    for i in range(len(obs)):
        acc_o += obs[i]; acc_e += exp[i]
        if acc_e >= 5:
            o.append(acc_o); e.append(acc_e); acc_o = acc_e = 0.0
    if acc_e > 0:
        if e:
            o[-1] += acc_o; e[-1] += acc_e
        else:
            o.append(acc_o); e.append(acc_e)
    o, e = np.array(o), np.array(e)
    e *= o.sum() / e.sum()                       # renormalizar
    chi2 = ((o - e) ** 2 / e).sum()
    dof = max(len(o) - 1 - 1, 1)                 # -1 por lambda estimado
    p = 1 - stats.chi2.cdf(chi2, dof)
    return dict(lam=float(lam), chi2=float(chi2), dof=dof, p=float(p),
                n_bins=len(o), n_horas=len(counts))


def fit_llegadas(data, evs):
    d = data[data > 0]
    # exponencial (proceso de Poisson): scale = media
    loc_e, scale_e = stats.expon.fit(d, floc=0)
    # gamma libre (para reportar la forma ~1)
    a, loc_g, scale_g = stats.gamma.fit(d, floc=0)
    hc, (lo, hi) = hourly_counts(evs)
    pois = poisson_chi2(hc)
    return dict(n=len(data), mean=float(data.mean()), std=float(data.std()),
                median=float(np.median(data)), min=float(data.min()),
                max=float(data.max()), exp_scale=float(scale_e),
                gamma_a=float(a), gamma_scale=float(scale_g),
                op_lo=lo, op_hi=hi, poisson=pois)


lleg_fit = {sc: fit_llegadas(d, events[sc]) for sc, d in llegadas.items()}


# ============================================================
# 3. RETIROS [SUPUESTO]  -> lognormal cola derecha
#    mediana ~600 min (10 h), p90 ~1 día, tail a ~2-3 días
# ============================================================
N = 50000
med_min = 600.0
sigma_ln = 0.70
mu_ln = np.log(med_min)
retiros = np.random.lognormal(mean=mu_ln, sigma=sigma_ln, size=N)
retiros = np.clip(retiros, 15, None)          # piso físico 15 min
retiros = np.round(retiros, 1)

# ============================================================
# 4. DEVOLUCIONES [SUPUESTO] -> inter-arribo, tasa anclada NRF 16.5%
#    Gamma asimétrica, media ~360 min (6 h entre devoluciones en el barrio)
# ============================================================
dev_mean = 360.0
dev_shape = 3.0                                # forma -> asimetría moderada
dev_scale = dev_mean / dev_shape
devoluciones = np.random.gamma(dev_shape, dev_scale, size=N)
devoluciones = np.clip(devoluciones, 5, None)
devoluciones = np.round(devoluciones, 1)


# ============================================================
# 5. ESCRITURA DE XLSX
# ============================================================
def write_col(path, header, values, extra_sheets=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append([header])
    for v in values:
        ws.append([float(v)])
    if extra_sheets:
        for name, rows in extra_sheets.items():
            s = wb.create_sheet(name)
            for row in rows:
                s.append(row)
    wb.save(path)


# 5.1 Llegadas reales por escenario (una muestra por archivo)
for sc, d in llegadas.items():
    f = f"{sc.capitalize()}" if sc != "CYBER" else "Cyber"
    write_col(os.path.join(OUT, f"dataset_llegadas_{sc.lower()}.xlsx"),
              "intervalo_arribo_minutos", d)

# 5.2 Tamaños: probabilidades por escenario, DOS mapeos
tam_probs = {"FULL": {}, "LOCK": {}}
wb = Workbook()
for key, mapping, title in [("FULL", TIER_FULL, "Completo"),
                            ("LOCK", TIER_LOCK, "Lockeable")]:
    ws = wb.create_sheet(title)
    ws.append(["escenario", "p_chico", "p_mediano", "p_grande",
               "n_chico", "n_mediano", "n_grande", "n_total"])
    for sc in ["NORMAL", "NAVIDAD", "CYBER"]:
        (pc, pm, pg), agg, tot = tier_probs(raw_size[sc], mapping)
        tam_probs[key][sc] = (pc, pm, pg)
        ws.append([sc, round(pc, 4), round(pm, 4), round(pg, 4),
                   agg["chico"], agg["mediano"], agg["grande"], tot])
# hoja con conteos crudos
wsr = wb.create_sheet("Crudo")
all_sizes = sorted({s for sc in raw_size for s in raw_size[sc]})
wsr.append(["escenario"] + all_sizes)
for sc in ["NORMAL", "NAVIDAD", "CYBER"]:
    wsr.append([sc] + [raw_size[sc].get(s, 0) for s in all_sizes])
del wb["Sheet"]
wb.save(os.path.join(OUT, "dataset_tamanos.xlsx"))

# 5.3 Retiros y devoluciones (supuestos)
write_col(os.path.join(OUT, "dataset_retiros.xlsx"),
          "Tiempo_Retiro_Minutos", retiros)
write_col(os.path.join(OUT, "dataset_devoluciones.xlsx"),
          "intervalo_arribo_minutos", devoluciones)


# ============================================================
# 6. RESUMEN EN CONSOLA
# ============================================================
print("=== LLEGADAS (REAL) - exponencial + Poisson(conteos/hora) ===")
for sc, ft in lleg_fit.items():
    p = ft["poisson"]
    print(f"{sc:8s} n={ft['n']:5d} media_IA={ft['mean']:5.2f}min gamma_a={ft['gamma_a']:.2f} "
          f"exp_scale={ft['exp_scale']:.2f} | Poisson lambda={p['lam']:.2f}/h "
          f"chi2={p['chi2']:.2f} dof={p['dof']} p={p['p']:.3f} (op {ft['op_lo']}-{ft['op_hi']}h)")

print("\n=== TAMAÑOS (REAL) - P(chico/mediano/grande) ===")
for key in ["FULL", "LOCK"]:
    print(f"  [{key}]")
    for sc, (pc, pm, pg) in tam_probs[key].items():
        print(f"    {sc:8s} chico={pc:.3f} mediano={pm:.3f} grande={pg:.3f}")

print("\n=== RETIROS (SUPUESTO lognormal) ===")
print(f"n={N} media={retiros.mean():.1f} mediana={np.median(retiros):.1f} "
      f"p90={np.percentile(retiros,90):.1f} p99={np.percentile(retiros,99):.1f} "
      f"max={retiros.max():.1f}")

print("\n=== DEVOLUCIONES (SUPUESTO gamma, ancla NRF 16.5%) ===")
print(f"n={N} media={devoluciones.mean():.1f} mediana={np.median(devoluciones):.1f} "
      f"p90={np.percentile(devoluciones,90):.1f} max={devoluciones.max():.1f}")

# export de parámetros para la documentación
import json
params = dict(llegadas=lleg_fit, tamanos=tam_probs, raw_size={k: dict(v) for k, v in raw_size.items()},
              retiros=dict(dist="lognormal", median_min=med_min, sigma=sigma_ln,
                           mean=float(retiros.mean()), p90=float(np.percentile(retiros,90)),
                           p99=float(np.percentile(retiros,99))),
              devoluciones=dict(dist="gamma", shape=dev_shape, scale=dev_scale,
                                mean=float(devoluciones.mean())))
with open(os.path.join(OUT, "_params.json"), "w") as f:
    json.dump(params, f, indent=2, default=float)
print("\nOK -> archivos escritos en outputs/")
