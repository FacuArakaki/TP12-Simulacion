# -*- coding: utf-8 -*-
"""Figuras para las SLIDES (a color, tamaño grande). Relee los xlsx de resultados."""
import os
from collections import defaultdict
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch, FancyArrowPatch
from openpyxl import load_workbook

D = os.path.dirname(os.path.abspath(__file__))
NAVY, AMBER, ICE, GREY = '#1E2761', '#E8833A', '#7FA6D9', '#5A6478'
plt.rcParams.update({'font.size': 12, 'axes.titlesize': 14, 'axes.labelsize': 12,
                     'xtick.labelsize': 11, 'ytick.labelsize': 11,
                     'legend.fontsize': 11, 'figure.dpi': 160,
                     'axes.spines.top': False, 'axes.spines.right': False})
SER = [NAVY, ICE, AMBER]
MK = ['o', 's', '^']


def col(f):
    ws = load_workbook(os.path.join(D, f), read_only=True)['Datos']
    return [r[0] for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]


# ---------------------------------------------------------- 1) ESQUEMA
fig, ax = plt.subplots(figsize=(11, 4.2))
ax.set_xlim(0, 11); ax.set_ylim(0, 4.2); ax.axis('off')

def caja(x, y, w, h, txt, fc, tc='white', fs=12, bold=True):
    ax.add_patch(FancyBboxPatch((x, y), w, h, boxstyle="round,pad=0.08,rounding_size=0.12",
                                fc=fc, ec='none'))
    ax.text(x + w/2, y + h/2, txt, ha='center', va='center', color=tc,
            fontsize=fs, fontweight='bold' if bold else 'normal', linespacing=1.4)

def flecha(x1, y1, x2, y2, c=GREY, txt=None, dy=0.18):
    ax.add_patch(FancyArrowPatch((x1, y1), (x2, y2), arrowstyle='-|>', mutation_scale=18,
                                 lw=2, color=c, shrinkA=0, shrinkB=0))
    if txt:
        ax.text((x1+x2)/2, max(y1, y2) + dy, txt, ha='center', va='bottom',
                fontsize=10.5, color=GREY)

caja(0.1, 2.45, 2.25, 1.0, "Entrega\ndomiciliaria\nfallida", NAVY, fs=11.5)
caja(4.15, 2.2, 2.7, 1.5, "RED DE LOCKERS\nchico · mediano · grande", AMBER, fs=12.5)
caja(8.65, 2.45, 2.25, 1.0, "Retiro del\nusuario", NAVY, fs=11.5)
caja(0.1, 0.35, 2.25, 0.95, "Devoluciones\n(usuario)", ICE, tc=NAVY, fs=11.5)
caja(8.65, 0.35, 2.25, 0.95, "Camión\ncada IPC", ICE, tc=NAVY, fs=11.5)

flecha(2.5, 2.95, 4.05, 2.95, NAVY, "por ausencia")
flecha(6.95, 2.95, 8.55, 2.95, NAVY, "retiro (TPR)")
flecha(2.5, 0.85, 4.6, 2.1, ICE)
ax.text(3.3, 1.35, "compiten por\nla misma capacidad", ha='center', va='center',
        fontsize=10.5, color=GREY, style='italic')
flecha(6.5, 2.1, 8.55, 0.85, ICE)
ax.text(7.9, 1.35, "retira devoluciones\ny vencidos (TMP)", ha='center', va='center',
        fontsize=10.5, color=GREY, style='italic')
ax.text(5.5, 3.95, "Política flexible: un paquete chico puede ocupar un compartimento mayor",
        ha='center', va='center', fontsize=12, color=NAVY, fontweight='bold')
plt.tight_layout(); plt.savefig(os.path.join(D, 'slide_esquema.png'), transparent=True); plt.close()

# ---------------------------------------------------------- 2) DATOS: llegadas reales
lleg = {e: col(f'dataset_llegadas_{e}.xlsx') for e in ('normal', 'navidad', 'cyber')}
fig, ax = plt.subplots(figsize=(7.2, 3.9))
for (e, d), c, m in zip(lleg.items(), SER, MK):
    dd = [x for x in d if x < 22]
    ax.hist(dd, bins=22, histtype='step', linewidth=2.6, color=c, density=True,
            label=f"{e.capitalize()} — {sum(d)/len(d):.2f} min")
ax.set_xlabel('Inter-arribo entre entregas fallidas (min)')
ax.set_ylabel('Densidad')
ax.legend(frameon=False)
plt.tight_layout(); plt.savefig(os.path.join(D, 'slide_datos.png'), transparent=True); plt.close()

# ---------------------------------------------------------- datos del barrido
ws = load_workbook(os.path.join(D, 'resultados_barrido.xlsx'), read_only=True)['Barrido']
cpe = defaultdict(dict); opt = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    cpe[r[0]][(r[1], r[2], r[3])] = r[5]
    if r[12] == 'SI':
        opt[r[0]] = (r[1], r[2], r[3])
GC = sorted({k[0] for k in cpe['CYBER']})

# ---------------------------------------------------------- 3) CURVA U
import json as _json
_optC = {e: opt[e][0] for e in opt}
fig, ax = plt.subplots(figsize=(7.6, 4.0))
for esc, c, m in zip(['NORMAL', 'NAVIDAD', 'CYBER'], SER, MK):
    _jf = os.path.join(D, f'curvaU_{esc}.json')
    if os.path.exists(_jf):
        _d = {int(k): v for k, v in _json.load(open(_jf)).items()}
        xs = [x for x in sorted(_d) if x >= 25]; ys = [_d[x] for x in xs]
    else:
        _, bm, bg = opt[esc]
        xs = [x for x in GC if (x, bm, bg) in cpe[esc]]
        ys = [cpe[esc][(x, bm, bg)] for x in xs]; _d = None
    ax.plot(xs, ys, marker=m, color=c, lw=2.8, ms=8, label=esc.capitalize())
    if _d and _optC[esc] in _d:
        ax.plot(_optC[esc], _d[_optC[esc]], marker='*', color=c, ms=18, mec='white', mew=1.2, zorder=5)
ax.annotate('zona óptima', xy=(50, 0.55), xytext=(80, 1.35),
            fontsize=12, color=AMBER, fontweight='bold',
            arrowprops=dict(arrowstyle='->', color=AMBER, lw=2))
ax.set_xlabel('Compartimentos chicos instalados')
ax.set_ylabel('Costo por paquete (USD)')
ax.legend(frameon=False); ax.grid(alpha=0.2)
plt.tight_layout(); plt.savefig(os.path.join(D, 'slide_curvaU.png'), transparent=True); plt.close()

# ---------------------------------------------------------- 4) COMPARACIÓN
ws = load_workbook(os.path.join(D, 'comparacion_configs.xlsx'), read_only=True)['Comparacion']
comp = defaultdict(dict)
for r in ws.iter_rows(min_row=2, values_only=True):
    comp[r[0]][r[1]] = (r[6], r[9])
CATS = ['Malo', 'Normal', 'Óptimo', 'Óptimo excelente']
NOM = ['Sub-\ndimensionado', 'Sin\noptimizar', 'ÓPTIMO', 'Máximo\nservicio']
fig, ax = plt.subplots(figsize=(7.6, 4.0))
xs = range(len(CATS)); w = 0.26
for i, (esc, c) in enumerate(zip(['NORMAL', 'NAVIDAD', 'CYBER'], SER)):
    ys = [comp[esc][k][0] for k in CATS]
    ax.bar([x + (i - 1) * w for x in xs], ys, w, color=c, label=esc.capitalize())
ax.set_yscale('log')
ax.set_xticks(list(xs)); ax.set_xticklabels(NOM, fontsize=11)
ax.set_ylabel('Costo por paquete (USD, escala log)')
ax.legend(frameon=False); ax.grid(axis='y', alpha=0.2)
plt.tight_layout(); plt.savefig(os.path.join(D, 'slide_comparacion.png'), transparent=True); plt.close()

# ---------------------------------------------------------- 5) TORNADO
ws = load_workbook(os.path.join(D, 'sensibilidad.xlsx'), read_only=True)['Tornado']
tor = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] == 'CYBER']
tor.sort(key=lambda r: abs(r[4] - r[3]))
base = tor[0][2]
NOMP = {'C_CAP (amortización lockers)': 'Amortización locker',
        'C_EF (entrega fallida)': 'Entrega fallida',
        'C_CAMION (pase de camión)': 'Pase de camión',
        'C_MANUAL (retiro manual)': 'Retiro manual',
        'C_VENCIDO (paq. vencido)': 'Paquete vencido',
        'Espacio ocioso': 'Espacio ocioso'}
fig, ax = plt.subplots(figsize=(7.6, 3.9))
for y, r in enumerate(tor):
    ax.barh(y, r[3] - r[2], left=r[2], color=ICE, height=0.62)
    ax.barh(y, r[4] - r[2], left=r[2], color=NAVY, height=0.62)
ax.axvline(base, color=AMBER, lw=2.4, ls='--')
ax.set_yticks(range(len(tor)))
ax.set_yticklabels([NOMP.get(r[1], r[1]) for r in tor])
ax.set_xlabel('Costo por paquete (USD)')
ax.legend(handles=[plt.Rectangle((0, 0), 1, 1, fc=ICE),
                   plt.Rectangle((0, 0), 1, 1, fc=NAVY)],
          labels=['costo −30 %', 'costo +30 %'], frameon=False, loc='lower right')
ax.grid(axis='x', alpha=0.2)
plt.tight_layout(); plt.savefig(os.path.join(D, 'slide_tornado.png'), transparent=True); plt.close()

print("OK -> slide_esquema / slide_datos / slide_curvaU / slide_comparacion / slide_tornado")
print("optimos:", opt)
