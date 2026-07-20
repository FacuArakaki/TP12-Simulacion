# -*- coding: utf-8 -*-
"""
============================================================================
 SIMULACIÓN EVENTO A EVENTO — Logística de última milla con red de lockers
 TP Simulación — UTN FRBA
============================================================================
 Motor de cátedra (event-scheduling a mano, HV = infinito). NO usa librerías
 de eventos. Consume los datasets generados (muestreo empírico / inverse-transform).

 Consolidación del Colab + integración de FDPs reales + función de costos +
 N réplicas con intervalo de confianza del 95%.

 CAMBIOS clave respecto del Colab original:
   - Llegadas = inter-arribo REAL de entregas fallidas por ausencia (dataset).
     -> Se ELIMINA el Random de ausencia (el dataset ya son fallos). Sin doble-conteo.
   - Distribución de tamaños REAL por escenario (reemplaza el 0.476/0.765 fijo).
   - Retiro = lognormal (dataset), no la lognormal de "cola pesada" equivocada.
   - Vencimiento: se detecta en el pase del camión (T - TOL >= TMP), consistente
     con el TEF implementado.
============================================================================
"""
import os
import math
import random
import statistics
from openpyxl import load_workbook

DATA_DIR = os.path.dirname(os.path.abspath(__file__))  # donde están los dataset_*.xlsx

# ===========================================================================
# 0. PARÁMETROS DE COSTO  ─── Anclados a benchmarks de industria (USD) ───
#    Detalle y fuentes en COSTOS_REFERENCIA.md.
#    [FUENTE]   = respaldado por estudio de industria citable.
#    [DEFINIDO] = definido por el equipo (sin fuente confiable) -> cubrir con sensibilidad.
#    Unidades: USD. Incertidumbre -> análisis de sensibilidad +-30%.
# ===========================================================================
# Amortización diaria por compartimento. Derivada de un TCO real a 10 años:
#   sistema outdoor de 30 compartimentos = EUR18.000 equipo + EUR2.000 sitio
#   + EUR20.000 operación = EUR40.000 / 10 años / 30 comp = EUR0.365 por comp-día
#   ~= USD0.395. Repartido por tamaño con el mix 43/43/16 y ratio 1 : 1.5 : 2.5.
# NOTA: el TCO no incluye alquiler del espacio que hospeda el locker [SUPUESTO no cubierto].
C_CAP_C = 0.27   # USD/día por locker chico     [FUENTE: ParcelHive, TCO 10 años]
C_CAP_M = 0.40   # USD/día por locker mediano   [FUENTE: idem]
C_CAP_G = 0.67   # USD/día por locker grande    [FUENTE: idem]
# Penalizaciones operativas:
C_EF        = 15.0   # USD por entrega fallida -> redespacho          [FUENTE: Optivo/GoBolt, $8 directo-$17.78 total]
C_MANUAL    = 25.0   # USD por devolución rechazada -> retiro manual  [FUENTE: Claimlane/Productiv, $15-30 por devolución]
C_VENCIDO   = 12.0   # USD por paquete vencido levantado por el camión [DEFINIDO: ~ proc. devolución, extremo bajo]
C_CAMION    = 8.0    # USD por pase del camión (combustible + chofer)  [DEFINIDO: orden de magnitud costo/parada last-mile]
# Costo por espacio ocioso (paquete en compartimento más grande que el suyo):
C_OCIOSO_CM = 0.5    # USD por chico-en-mediano  [DEFINIDO: costo de oportunidad de capacidad bloqueada]
C_OCIOSO_CG = 1.0    # USD por chico-en-grande   [DEFINIDO]
C_OCIOSO_MG = 0.7    # USD por mediano-en-grande [DEFINIDO]

# ===========================================================================
# 1. VARIABLES DE CONTROL (lo que se optimiza en el barrido)
# ===========================================================================
DEFAULT_CONFIG = dict(CLTC=20, CLTM=15, CLTG=10)  # cantidad de lockers por tamaño
TMP = 60 * 24 * 3          # Tiempo Máximo de Permanencia (3 días, en min)
IPC = 60 * 24              # Intervalo de Pase del Camión (1 vez/día, en min)
TF  = 60 * 24 * 365        # Horizonte de simulación (1 año, en min)

# Escala BARRIO: la data de llegadas es de TODA la operación (~5 zonas).
# El modelo es 1 punto de lockers = 1 barrio -> se estira el inter-arribo.
# SCALE=5 ≈ una zona. [SUPUESTO / decisión de alcance a validar con el grupo]
SCALE_BARRIO = 5.0

HV = float('inf')          # High Value (locker libre / evento inexistente)

# ===========================================================================
# 2. CARGA DE DATASETS  (las FDPs viven en los .xlsx generados)
# ===========================================================================
def _load_col(fname):
    ws = load_workbook(os.path.join(DATA_DIR, fname), read_only=True)['Datos']
    return [r[0] for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]

def _load_tamanos(mapeo='Lockeable'):
    ws = load_workbook(os.path.join(DATA_DIR, 'dataset_tamanos.xlsx'), read_only=True)[mapeo]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # escenario -> (p_chico, p_mediano, p_grande)
    return {r[0]: (float(r[1]), float(r[2]), float(r[3])) for r in rows}

LLEGADAS = {
    'NORMAL':  _load_col('dataset_llegadas_normal.xlsx'),
    'NAVIDAD': _load_col('dataset_llegadas_navidad.xlsx'),
    'CYBER':   _load_col('dataset_llegadas_cyber.xlsx'),
}
RETIROS      = _load_col('dataset_retiros.xlsx')
DEVOLUCIONES = _load_col('dataset_devoluciones.xlsx')
TAMANOS      = _load_tamanos('Lockeable')   # DEFAULT: mapeo lockeable (realista)

# ===========================================================================
# 3. MUESTREADORES (inverse-transform empírico sobre los datasets)
# ===========================================================================
class Sampler:
    """Muestreo empírico con STREAMS SEPARADOS por proceso estocástico.
    Números Aleatorios Comunes (CRN): dos configuraciones con la misma semilla
    ven la MISMA secuencia de llegadas, tamaños, retiros y devoluciones, sin
    importar cómo diverjan -> la diferencia de resultado es solo por los lockers,
    no por ruido. Reduce la varianza al comparar configuraciones en el barrido."""
    def __init__(self, seed):
        # Streams separados: la semilla se combina como str (py3.13 ya no
        # acepta tuplas como seed, solo None/int/float/str/bytes/bytearray).
        self.rng_lleg = random.Random(f"{seed}-1")
        self.rng_tam  = random.Random(f"{seed}-2")
        self.rng_ret  = random.Random(f"{seed}-3")
        self.rng_dev  = random.Random(f"{seed}-4")
    def llegada(self, arr):
        return arr[self.rng_lleg.randrange(len(arr))]
    def retiro(self, arr):
        return arr[self.rng_ret.randrange(len(arr))]
    def devolucion(self, arr):
        return arr[self.rng_dev.randrange(len(arr))]
    def tamano(self, probs):
        # 1=chico, 2=mediano, 3=grande
        u = self.rng_tam.random()
        pc, pm, pg = probs
        if u < pc: return 1
        if u < pc + pm: return 2
        return 3

# ===========================================================================
# 3b. FUNCIÓN DE COSTOS (separada del motor)
#     Permite recalcular el costo con otros parámetros SIN volver a simular,
#     porque los costos no afectan la dinámica -> sensibilidad exacta y barata.
# ===========================================================================
def costos_default():
    """Diccionario con los costos vigentes del módulo."""
    return dict(CAP_C=C_CAP_C, CAP_M=C_CAP_M, CAP_G=C_CAP_G, EF=C_EF,
                MANUAL=C_MANUAL, VENCIDO=C_VENCIDO, CAMION=C_CAMION,
                OCIOSO_CM=C_OCIOSO_CM, OCIOSO_CG=C_OCIOSO_CG, OCIOSO_MG=C_OCIOSO_MG)


def calcular_costos(m, C=None):
    """Costo total y CPE a partir de los CONTADORES CRUDOS de una corrida.
    m = contadores devueltos por simular(); C = dict de costos (default: el del módulo).
    Devuelve (costo_total, CPE)."""
    if C is None:
        C = costos_default()
    costo_cap = (m['CLTC'] * C['CAP_C'] + m['CLTM'] * C['CAP_M']
                 + m['CLTG'] * C['CAP_G']) * m['dias']
    costo_ef      = m['n_EF']       * C['EF']
    costo_manual  = m['n_CI']       * C['MANUAL']
    costo_venc    = m['n_vencidos'] * C['VENCIDO']
    costo_camion  = m['n_pases']    * C['CAMION']
    costo_ocioso  = (m['n_CEnM'] * C['OCIOSO_CM'] + m['n_CEnG'] * C['OCIOSO_CG']
                     + m['n_MEnG'] * C['OCIOSO_MG'])
    total = (costo_cap + costo_ef + costo_manual + costo_venc
             + costo_camion + costo_ocioso)
    return total, total / max(m['n_entregados'], 1)


# ===========================================================================
# 4. MOTOR DE SIMULACIÓN (una corrida = una réplica)
# ===========================================================================
def simular(escenario, config, seed):
    s = Sampler(seed)
    CLTC, CLTM, CLTG = config['CLTC'], config['CLTM'], config['CLTG']
    probs = TAMANOS[escenario]
    ia_llegadas = LLEGADAS[escenario]

    # --- Estado de lockers: HV=libre, 1=normal, 2=devolución ---
    DispC = [HV] * CLTC; DispM = [HV] * CLTM; DispG = [HV] * CLTG
    TOLC  = [HV] * CLTC; TOLM  = [HV] * CLTM; TOLG  = [HV] * CLTG   # tiempo de ocupación
    TPRC  = [HV] * CLTC; TPRM  = [HV] * CLTM; TPRG  = [HV] * CLTG   # tiempo próx. retiro

    # --- Contadores de resultado ---
    R = dict(CP=0, entregados=0, EF=0, dev_total=0, CI=0, vencidos=0,
             CantCEnM=0, CantCEnG=0, CantMEnG=0, pases=0)

    def colocar(size, op, T, t_ret):
        """Política flexible: chico->[C,M,G]; mediano->[M,G]; grande->[G].
           t_ret = tiempo de retiro pre-sampleado (solo se usa si op==1).
           Devuelve True si se colocó, False si saturación."""
        orden = {1: [('C',), ('M',), ('G',)], 2: [('M',), ('G',)], 3: [('G',)]}[size]
        for (tier,) in orden:
            disp, tol, tpr, cap = {
                'C': (DispC, TOLC, TPRC, CLTC),
                'M': (DispM, TOLM, TPRM, CLTM),
                'G': (DispG, TOLG, TPRG, CLTG),
            }[tier]
            for i in range(cap):
                if disp[i] == HV:
                    disp[i] = op
                    tol[i] = T
                    # ocioso si el paquete fue a un compartimento mayor
                    if size == 1 and tier == 'M': R['CantCEnM'] += 1
                    if size == 1 and tier == 'G': R['CantCEnG'] += 1
                    if size == 2 and tier == 'G': R['CantMEnG'] += 1
                    # retiro por el usuario solo si es pedido normal (op==1)
                    if op == 1:
                        tpr[i] = T + t_ret
                    return True
        return False

    def retirar_devoluciones():
        for disp, tpr in ((DispC, TPRC), (DispM, TPRM), (DispG, TPRG)):
            for i in range(len(disp)):
                if disp[i] == 2:
                    disp[i] = HV; tpr[i] = HV

    def retirar_vencidos(T):
        n = 0
        for disp, tol, tpr in ((DispC, TOLC, TPRC), (DispM, TOLM, TPRM), (DispG, TOLG, TPRG)):
            for i in range(len(disp)):
                if disp[i] == 1 and (T - tol[i]) >= TMP:
                    disp[i] = HV; tol[i] = HV; tpr[i] = HV; n += 1
        return n

    # --- Calendario de eventos (TEF) ---
    T = 0.0
    TPLLP = s.llegada(ia_llegadas) * SCALE_BARRIO  # próx. llegada (fallo por ausencia)
    TPPC  = IPC                                    # próx. pase de camión
    TPDPC = s.devolucion(DEVOLUCIONES)             # próx. devolución chica
    TPDPM = s.devolucion(DEVOLUCIONES)
    TPDPG = s.devolucion(DEVOLUCIONES)

    while True:
        min_RC = min(TPRC) if TPRC else HV
        min_RM = min(TPRM) if TPRM else HV
        min_RG = min(TPRG) if TPRG else HV
        eventos = {'LLP': TPLLP, 'PC': TPPC, 'DPC': TPDPC, 'DPM': TPDPM,
                   'DPG': TPDPG, 'RC': min_RC, 'RM': min_RM, 'RG': min_RG}
        ev = min(eventos, key=eventos.get)
        T = eventos[ev]
        if T > TF:
            break

        if ev == 'LLP':                          # llegada de pedido normal al locker
            R['CP'] += 1
            TPLLP = T + s.llegada(ia_llegadas) * SCALE_BARRIO
            size = s.tamano(probs)
            t_ret = s.retiro(RETIROS)            # se consume siempre -> mantiene CRN
            if not colocar(size, 1, T, t_ret):
                R['EF'] += 1                     # locker saturado -> entrega fallida

        elif ev == 'PC':                         # pase del camión
            TPPC = T + IPC
            R['pases'] += 1
            retirar_devoluciones()
            R['vencidos'] += retirar_vencidos(T)

        elif ev in ('DPC', 'DPM', 'DPG'):        # llegada de devolución
            size = {'DPC': 1, 'DPM': 2, 'DPG': 3}[ev]
            R['dev_total'] += 1
            if ev == 'DPC': TPDPC = T + s.devolucion(DEVOLUCIONES)
            elif ev == 'DPM': TPDPM = T + s.devolucion(DEVOLUCIONES)
            else: TPDPG = T + s.devolucion(DEVOLUCIONES)
            if not colocar(size, 2, T, 0.0):
                R['CI'] += 1                     # sin lugar -> cliente insatisfecho + retiro manual

        else:                                    # retiro por el usuario (RC/RM/RG)
            disp, tpr = {'RC': (DispC, TPRC), 'RM': (DispM, TPRM), 'RG': (DispG, TPRG)}[ev]
            idx = tpr.index({'RC': min_RC, 'RM': min_RM, 'RG': min_RG}[ev])
            disp[idx] = HV; tpr[idx] = HV
            R['entregados'] += 1

    # --- Contadores crudos + costos ---
    # Los contadores se devuelven para poder RECALCULAR costos sin re-simular
    # (los parámetros de costo no afectan la dinámica de la simulación).
    m = dict(CLTC=CLTC, CLTM=CLTM, CLTG=CLTG, dias=TF / (60 * 24),
             n_EF=R['EF'], n_CI=R['CI'], n_vencidos=R['vencidos'],
             n_pases=R['pases'], n_CEnM=R['CantCEnM'], n_CEnG=R['CantCEnG'],
             n_MEnG=R['CantMEnG'], n_entregados=R['entregados'])
    costo_total, cpe = calcular_costos(m)

    total_norm = max(R['CP'], 1)
    out = dict(
        CostoTotal=costo_total,
        CPE=cpe,                                            # $/paquete entregado
        entregados=R['entregados'],
        EF_pct=100 * R['EF'] / total_norm,                  # % entregas fallidas (saturación normal)
        CI=R['CI'],                                         # clientes insatisfechos (dev. rechazadas)
        PPV=100 * R['vencidos'] / total_norm,               # % paquetes vencidos
        EO=R['CantCEnM'] + R['CantCEnG'] + R['CantMEnG'],   # unidades de espacio ocioso
        vencidos=R['vencidos'],
        CP=R['CP'], dev_total=R['dev_total'],
    )
    out.update(m)          # contadores crudos, para el análisis de sensibilidad
    return out

# ===========================================================================
# 5. N RÉPLICAS + INTERVALO DE CONFIANZA 95%
# ===========================================================================
def replicar(escenario, config, n_rep=30, seed0=1000):
    corridas = [simular(escenario, config, seed0 + k) for k in range(n_rep)]
    out = {}
    for key in corridas[0]:
        vals = [c[key] for c in corridas]
        m = statistics.mean(vals)
        sd = statistics.stdev(vals) if len(vals) > 1 else 0.0
        # IC 95% (t de Student aprox 1.96 para n>=30; usar t exacto si n<30)
        tcrit = 2.045 if n_rep < 30 else 1.96
        hw = tcrit * sd / math.sqrt(n_rep)     # semiancho del IC
        out[key] = (m, hw)
    return out

# ===========================================================================
# 6. MAIN — corrida de prueba con la config por defecto
# ===========================================================================
if __name__ == '__main__':
    config = DEFAULT_CONFIG
    N_REP = 30
    print(f"Config lockers: {config} | TMP={TMP/1440:.0f}d IPC={IPC/1440:.0f}d "
          f"TF={TF/1440:.0f}d | réplicas={N_REP} | mapeo tamaños=Lockeable\n")
    header = f"{'Escenario':<9} {'CPE ($/paq)':>16} {'Costo Total ($)':>20} {'EF %':>10} {'PPV %':>10} {'CI':>10} {'EO':>8}"
    print(header); print("-" * len(header))
    for esc in ['NORMAL', 'NAVIDAD', 'CYBER']:
        r = replicar(esc, config, n_rep=N_REP)
        cpe_m, cpe_h = r['CPE']
        ct_m, ct_h = r['CostoTotal']
        ef_m, ef_h = r['EF_pct']
        ppv_m, ppv_h = r['PPV']
        ci_m, ci_h = r['CI']
        eo_m, eo_h = r['EO']
        print(f"{esc:<9} {cpe_m:8.1f}±{cpe_h:5.1f} {ct_m:13.0f}±{ct_h:5.0f} "
              f"{ef_m:6.2f}±{ef_h:3.2f} {ppv_m:6.2f}±{ppv_h:3.2f} "
              f"{ci_m:6.0f}±{ci_h:3.0f} {eo_m:5.0f}±{eo_h:2.0f}")
    print("\n(± = semiancho del intervalo de confianza al 95%)")
