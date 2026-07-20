# -*- coding: utf-8 -*-
"""
============================================================================
 ANÁLISIS DE SENSIBILIDAD — COSTOS (Grupo A)
 TP Simulación — lockers última milla
============================================================================
 Varía cada parámetro de costo ±30% (OAT, uno por vez) y mide:

   1) TORNADO   -> cuánto se mueve el CPE con cada costo (ranking de impacto).
   2) ROBUSTEZ  -> si el DISEÑO ÓPTIMO (CLTC/CLTM/CLTG) cambia con esa variación.

 Truco clave: los costos NO afectan la dinámica de la simulación, solo el cálculo
 final. Entonces se simula UNA vez, se guardan los contadores crudos, y cada
 variación de costo se recalcula con S.calcular_costos() -> exacto y sin ruido
 (no hay diferencia de semillas entre variantes).

 Requiere: simulacion_lockers.py + dataset_*.xlsx (+ resultados_barrido.xlsx
 para tomar el óptimo; si no está, usa DEFAULT_CONFIG).

 Genera: sensibilidad.xlsx, sensibilidad_tornado.png

 Corrida rápida (para probar):
   SENS_TF_DIAS=30 SENS_NREP=15 SENS_NREP_GRID=6 python3 sensibilidad.py
============================================================================
"""
import os
import statistics as st
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

import simulacion_lockers as S

# --- Configuración (overrideable por env para pruebas rápidas) ---
TF_DIAS = int(os.environ.get('SENS_TF_DIAS', 0))
if TF_DIAS:
    S.TF = 60 * 24 * TF_DIAS
N_REP_BASE = int(os.environ.get('SENS_NREP', 30))        # tornado (config óptima)
N_REP_GRID = int(os.environ.get('SENS_NREP_GRID', 10))   # robustez (grilla completa)
SEED0 = 1000
DELTA = 0.30

GRID_CLTC = [10, 20, 30, 40, 50, 60]
GRID_CLTM = [5, 10, 15, 20, 25]
GRID_CLTG = [5, 10, 15]
ESCENARIOS = ['NORMAL', 'NAVIDAD', 'CYBER']

# Bloques de costo: se varían juntos los que representan un mismo concepto
BLOQUES = {
    'C_EF (entrega fallida)':       ['EF'],
    'C_MANUAL (retiro manual)':     ['MANUAL'],
    'C_CAP (amortización lockers)': ['CAP_C', 'CAP_M', 'CAP_G'],
    'C_VENCIDO (paq. vencido)':     ['VENCIDO'],
    'C_CAMION (pase de camión)':    ['CAMION'],
    'Espacio ocioso':               ['OCIOSO_CM', 'OCIOSO_CG', 'OCIOSO_MG'],
}


def costos_con(bloque, factor):
    """Costos por defecto con un bloque escalado por 'factor'."""
    C = S.costos_default()
    for k in BLOQUES[bloque]:
        C[k] *= factor
    return C


def optimos_desde_barrido():
    """Lee la config óptima por escenario de resultados_barrido.xlsx."""
    try:
        ws = load_workbook('resultados_barrido.xlsx', read_only=True)['Barrido']
        out = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[12] == 'SI':                     # columna 'optimo'
                out[r[0]] = (r[1], r[2], r[3])
        return out
    except Exception as e:
        print(f"  (no se pudo leer resultados_barrido.xlsx: {e})")
        return {}


def cpe_medio(corridas, C):
    """CPE promedio de un set de corridas, recalculado con los costos C."""
    return st.mean(S.calcular_costos(m, C)[1] for m in corridas)


# ============================================================
# 1. TORNADO — impacto de cada costo sobre el CPE
# ============================================================
opt = optimos_desde_barrido()
if not opt:
    d = S.DEFAULT_CONFIG
    opt = {e: (d['CLTC'], d['CLTM'], d['CLTG']) for e in ESCENARIOS}
    print("  -> usando DEFAULT_CONFIG como base")

print(f"Sensibilidad de costos (+-{DELTA:.0%}) | TF={S.TF/1440:.0f}d | "
      f"réplicas tornado={N_REP_BASE}, grilla={N_REP_GRID}\n")

tornado = {}     # esc -> [(bloque, cpe_base, cpe_low, cpe_high)]
for esc in ESCENARIOS:
    c, m_, g = opt[esc]
    cfg = dict(CLTC=c, CLTM=m_, CLTG=g)
    corridas = [S.simular(esc, cfg, SEED0 + k) for k in range(N_REP_BASE)]
    base = cpe_medio(corridas, S.costos_default())
    filas = []
    for bloque in BLOQUES:
        lo = cpe_medio(corridas, costos_con(bloque, 1 - DELTA))
        hi = cpe_medio(corridas, costos_con(bloque, 1 + DELTA))
        filas.append((bloque, base, lo, hi))
    filas.sort(key=lambda t: abs(t[3] - t[2]), reverse=True)
    tornado[esc] = filas

    print(f"===== {esc}  (óptimo {c}/{m_}/{g}, CPE base = ${base:.2f}) =====")
    print(f"{'Parámetro':<32}{'-30%':>9}{'+30%':>9}{'variación':>12}")
    for bloque, b, lo, hi in filas:
        var = 100 * (hi - lo) / b
        print(f"{bloque:<32}{lo:>9.2f}{hi:>9.2f}{var:>11.1f}%")
    print()


# ============================================================
# 2. ROBUSTEZ DEL ÓPTIMO
#    Se simula la grilla UNA vez y se re-optimiza con cada costo variado.
# ============================================================
print("Robustez del óptimo (¿cambia el diseño ganador?)...\n")
robustez = {}    # esc -> (config_base, [(bloque, lbl, config, cambia)])
for esc in ESCENARIOS:
    grid = {}
    for c in GRID_CLTC:
        for m_ in GRID_CLTM:
            for g in GRID_CLTG:
                cfg = dict(CLTC=c, CLTM=m_, CLTG=g)
                grid[(c, m_, g)] = [S.simular(esc, cfg, SEED0 + k)
                                    for k in range(N_REP_GRID)]

    def mejor_con(C, _grid=grid):
        return min(_grid.items(), key=lambda kv: cpe_medio(kv[1], C))[0]

    base_cfg = mejor_con(S.costos_default())
    filas = []
    for bloque in BLOQUES:
        for factor, lbl in ((1 - DELTA, '-30%'), (1 + DELTA, '+30%')):
            cfg_v = mejor_con(costos_con(bloque, factor))
            filas.append((bloque, lbl, cfg_v, cfg_v != base_cfg))
    robustez[esc] = (base_cfg, filas)

    cambios = sum(1 for f in filas if f[3])
    print(f"  {esc:8s} óptimo base = {base_cfg[0]}/{base_cfg[1]}/{base_cfg[2]} "
          f"| variaciones que CAMBIAN el óptimo: {cambios}/{len(filas)}")
    for bloque, lbl, cfg_v, cambia in filas:
        if cambia:
            print(f"      · {bloque} {lbl} -> {cfg_v[0]}/{cfg_v[1]}/{cfg_v[2]}")
print()


# ============================================================
# 3. EXPORTAR XLSX
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "Tornado"
ws.append(["escenario", "parametro", "CPE_base", "CPE_-30%", "CPE_+30%", "variacion_%"])
for esc in ESCENARIOS:
    for bloque, b, lo, hi in tornado[esc]:
        ws.append([esc, bloque, round(b, 3), round(lo, 3), round(hi, 3),
                   round(100 * (hi - lo) / b, 2)])

ws2 = wb.create_sheet("Robustez")
ws2.append(["escenario", "optimo_base", "parametro", "variacion",
            "optimo_resultante", "cambia"])
for esc in ESCENARIOS:
    base_cfg, filas = robustez[esc]
    for bloque, lbl, cfg_v, cambia in filas:
        ws2.append([esc, f"{base_cfg[0]}/{base_cfg[1]}/{base_cfg[2]}", bloque, lbl,
                    f"{cfg_v[0]}/{cfg_v[1]}/{cfg_v[2]}", "SI" if cambia else "no"])
wb.save("sensibilidad.xlsx")


# ============================================================
# 4. GRÁFICO — Tornado por escenario
# ============================================================
fig, axs = plt.subplots(1, 3, figsize=(19, 5.5))
for ax, esc in zip(axs, ESCENARIOS):
    filas = tornado[esc][::-1]          # el de mayor impacto queda arriba
    labels = [f[0] for f in filas]
    base = filas[0][1]
    ys = list(range(len(filas)))
    for y, (bloque, b, lo, hi) in zip(ys, filas):
        ax.barh(y, lo - b, left=b, color='#4C78A8', height=0.6)   # -30%
        ax.barh(y, hi - b, left=b, color='#E45756', height=0.6)   # +30%
    ax.axvline(base, color='black', lw=1.2, ls='--')
    ax.set_yticks(ys); ax.set_yticklabels(labels, fontsize=8)
    ax.set_xlabel("CPE ($/paquete entregado)")
    ax.set_title(f"{esc}  (CPE base = ${base:.2f})")
    ax.grid(axis='x', alpha=0.3)
handles = [plt.Rectangle((0, 0), 1, 1, color='#4C78A8'),
           plt.Rectangle((0, 0), 1, 1, color='#E45756')]
fig.legend(handles, ['costo -30%', 'costo +30%'], loc='lower center', ncol=2)
fig.suptitle("Sensibilidad de costos: impacto sobre el CPE (línea punteada = caso base)",
             fontsize=13)
plt.tight_layout(rect=[0, 0.06, 1, 1])
plt.savefig("sensibilidad_tornado.png", dpi=110)

print("OK -> sensibilidad.xlsx, sensibilidad_tornado.png")
